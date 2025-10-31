import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('user_input_files/financial data.xlsx', data_only=True)

# Extract all sheets
data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_data = []
    for row in ws.iter_rows(values_only=True):
        sheet_data.append(row)
    data[sheet_name] = sheet_data

# Save to JSON for easier processing
with open('tmp/financial_data.json', 'w') as f:
    json.dump(data, f, indent=2, default=str)

print(f"Extracted {len(data)} sheets: {list(data.keys())}")
print("\nFirst sheet preview:")
if data:
    first_sheet = list(data.values())[0]
    for i, row in enumerate(first_sheet[:10]):
        print(f"Row {i}: {row}")
